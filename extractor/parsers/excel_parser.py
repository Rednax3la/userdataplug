"""
Excel Parser — handles both .xls (xlrd) and .xlsx (openpyxl).
Returns structured table blocks where possible.
"""

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

MAX_ROWS = 50_000


def parse_excel(file_path: str) -> list[dict[str, Any]]:
    """Parse XLS/XLSX and return table blocks."""
    suffix = Path(file_path).suffix.lower()
    if suffix == ".xls":
        return _parse_xls(file_path)
    else:
        return _parse_xlsx(file_path)


def _parse_xlsx(file_path: str) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed")
        return []

    blocks = []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True, max_row=MAX_ROWS):
                clean = [str(c).strip() if c is not None else "" for c in row]
                if any(clean):  # skip fully empty rows
                    rows.append(clean)

            if rows:
                blocks.append({
                    "type": "table",
                    "content": rows,
                    "sheet": sheet_name,
                    "source": "xlsx",
                })
        wb.close()
    except Exception as e:
        logger.error(f"XLSX parse error: {e}")

    return blocks


def _parse_xls(file_path: str) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError:
        logger.error("xlrd not installed")
        return []

    blocks = []
    try:
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
            rows = []
            for row_idx in range(min(sheet.nrows, MAX_ROWS)):
                row = [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
                if any(row):
                    rows.append(row)

            if rows:
                blocks.append({
                    "type": "table",
                    "content": rows,
                    "sheet": sheet.name,
                    "source": "xls",
                })
    except Exception as e:
        logger.error(f"XLS parse error: {e}")

    return blocks
